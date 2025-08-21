# ────────────────────────── raporty_bot.py
# Jednowiadomościowy „panel” jak u BotFather – wersja poprawiona przez senior dev:
# ✔️ Widoczny raport dnia już na /start (jeśli istnieje) + blokada tworzenia nowego przy istniejącym
# ✔️ Kropki (●) przy „Zadania” i „Uwagi” oraz w samym panelu – jasna informacja o aktywnym trybie wprowadzania tekstu
# ✔️ „Zapisz do Excela” automatycznie doda bieżącą pozycję (jeśli kompletna), nie wymaga wciskania „Dodaj pozycję”
# ✔️ Pełna weryfikacja nakładania godzin (z pozycjami w panelu i zapisanymi w Excelu)
# ✔️ Time picker: 00 minut domyślnie, w edycji pre-selekcja aktualnej wartości
# ✔️ Emoji w panelu edycji, możliwość dodania nowej pozycji z panelu edycji
# ✔️ Presety miejsc (pobierane prawidłowo per użytkownik)
# ✔️ Eksport z przycisków i komend – nie blokuje UI
# ✔️ Locki, backupy, opcjonalny upload do SharePoint
# ⚠️ Trwałość danych: ustaw DATA_DIR=/data i podłącz wolumen w Northflank

import os
import re
import json
import logging
import shutil
import calendar as cal
from dataclasses import dataclass
from datetime import datetime, date, timedelta
from typing import Dict, List, Optional, Tuple

from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# ───────────── SharePoint (opcjonalny upload) ─────────────
try:
    from office365.sharepoint.client_context import ClientContext
    from office365.runtime.auth.client_credential import ClientCredential
except ModuleNotFoundError:
    ClientContext = ClientCredential = None  # brak biblioteki → upload pomijamy

# ───────────── Telegram ─────────────
from telegram import (
    Update,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    BotCommand,
)
from telegram.ext import (
    ApplicationBuilder,
    Application,
    CommandHandler,
    CallbackQueryHandler,
    MessageHandler,
    ContextTypes,
    ConversationHandler,
    filters,
)
from telegram.error import BadRequest

# ───────────── File locking & atomic save ─────────────
import tempfile
import portalocker

# ──────────────────── konfiguracja ────────────────────
load_dotenv()

TELEGRAM_TOKEN = os.getenv("TELEGRAM_TOKEN")  # MUST HAVE
WEBHOOK_URL = os.getenv("WEBHOOK_URL", "").rstrip("/")  # np. https://app-xyz.northflank.app
PORT = int(os.getenv("PORT", 8080))  # Northflank zwykle 8080

# ⚠️ Podłącz trwały wolumen w Northflank i ustaw DATA_DIR=/data
DATA_DIR = os.getenv("DATA_DIR", ".")
os.makedirs(DATA_DIR, exist_ok=True)
BACKUP_DIR = os.path.join(DATA_DIR, "backups")
os.makedirs(BACKUP_DIR, exist_ok=True)
BACKUP_KEEP = int(os.getenv("BACKUP_KEEP", "20"))

# opcjonalne SharePoint
SHAREPOINT_SITE = os.getenv("SHAREPOINT_SITE")
SHAREPOINT_DOC_LIB = os.getenv("SHAREPOINT_DOC_LIB")
SHAREPOINT_CLIENT_ID = os.getenv("SHAREPOINT_CLIENT_ID")
SHAREPOINT_CLIENT_SECRET = os.getenv("SHAREPOINT_CLIENT_SECRET")

EXCEL_FILE = os.path.join(DATA_DIR, "reports.xlsx")
MAPPING_FILE = os.path.join(DATA_DIR, "report_msgs.json")
PRESETS_FILE = os.path.join(DATA_DIR, "presets.json")
LOCK_FILE = os.path.join(DATA_DIR, "reports.lock")

ADMIN_IDS = {int(x) for x in os.getenv("ADMIN_IDS", "").split(",") if x.strip().isdigit()}

# ──────────────────── stałe excela ────────────────────
HEADERS = [
    "ID",        # {user_id}_{dd.mm.YYYY}_{idx}
    "Data",
    "Imię",
    "Miejsce",
    "Start",
    "Koniec",
    "Zadania",
    "Uwagi",
]
COLS = {name: i + 1 for i, name in enumerate(HEADERS)}  # 1-based

# ──────────────────── stany ────────────────────
DATE_PICK = 10
OVERLAP_DECIDE = 11

# ──────────────────── helpers: excel/lock/backup ────────────────────
def _atomic_save_wb(wb: Workbook, path: str) -> None:
    fd, tmp_path = tempfile.mkstemp(dir=os.path.dirname(path), suffix=".tmp")
    os.close(fd)
    wb.save(tmp_path)
    os.replace(tmp_path, path)

def _with_lock(fn, *args, **kwargs):
    with portalocker.Lock(LOCK_FILE, timeout=30):
        return fn(*args, **kwargs)

def _backup_file():
    if not os.path.exists(EXCEL_FILE):
        return
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    dst = os.path.join(BACKUP_DIR, f"reports_{ts}.xlsx")
    try:
        shutil.copy2(EXCEL_FILE, dst)
    except Exception as e:
        logging.warning("Backup failed: %s", e)
    files = sorted([f for f in os.listdir(BACKUP_DIR) if f.startswith("reports_") and f.endswith(".xlsx")])
    if len(files) > BACKUP_KEEP:
        for old in files[: len(files) - BACKUP_KEEP]:
            try:
                os.remove(os.path.join(BACKUP_DIR, old))
            except Exception:
                pass

def open_wb() -> Workbook:
    if os.path.exists(EXCEL_FILE):
        return load_workbook(EXCEL_FILE)
    return Workbook()

def month_key_from_date(date_str: str) -> str:
    d = datetime.strptime(date_str, "%d.%m.%Y")
    return f"{d.year:04d}-{d.month:02d}"

def ensure_month_sheet(wb: Workbook, month_key: str) -> Worksheet:
    ws: Optional[Worksheet] = wb[month_key] if month_key in wb.sheetnames else None
    if ws is None:
        ws = wb.create_sheet(title=month_key, index=0)
        ws.append(HEADERS)
        if "Sheet" in wb.sheetnames and wb["Sheet"].max_row == 1 and wb["Sheet"].max_column == 1:
            wb.remove(wb["Sheet"])
    else:
        idx = wb.sheetnames.index(month_key)
        if idx != 0:
            wb.move_sheet(ws, offset=-idx)
    return ws

def get_month_sheet_if_exists(wb: Workbook, month_key: str) -> Optional[Worksheet]:
    return wb[month_key] if month_key in wb.sheetnames else None

def report_exists(user_id: int, date_str: str) -> bool:
    if not os.path.exists(EXCEL_FILE):
        return False
    def _exists():
        wb = load_workbook(EXCEL_FILE)
        ws = get_month_sheet_if_exists(wb, month_key_from_date(date_str))
        if not ws:
            return False
        prefix = f"{user_id}_{date_str}_"
        for row in ws.iter_rows(min_row=2, values_only=True):
            if (row and row[0]) and str(row[0]).startswith(prefix):
                return True
        return False
    return _with_lock(_exists)

def save_report(entries: List[Dict[str, str]], user_id: int, date_str: str, name: str) -> None:
    def _save():
        wb = open_wb()
        ws = ensure_month_sheet(wb, month_key_from_date(date_str))
        prefix = f"{user_id}_{date_str}_"
        existing_idxs: List[int] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            rid = str(row[0]) if row and row[0] is not None else ""
            if rid.startswith(prefix):
                try:
                    existing_idxs.append(int(rid.split("_")[-1]))
                except Exception:
                    pass
        next_idx = (max(existing_idxs) + 1) if existing_idxs else 1
        for off, e in enumerate(entries):
            idx = next_idx + off
            ws.append([
                f"{user_id}_{date_str}_{idx}",
                date_str,
                name,
                e.get("place", ""),
                e.get("start", ""),
                e.get("end", ""),
                e.get("tasks", ""),
                e.get("notes", ""),
            ])
        _backup_file()
        _atomic_save_wb(wb, EXCEL_FILE)
    _with_lock(_save)
    _maybe_upload_sharepoint()

def read_entries_for_day(user_id: int, date_str: str) -> List[Dict[str, str]]:
    if not os.path.exists(EXCEL_FILE):
        return []
    def _read():
        wb = load_workbook(EXCEL_FILE)
        ws = get_month_sheet_if_exists(wb, month_key_from_date(date_str))
        if not ws:
            return []
        prefix = f"{user_id}_{date_str}_"
        out: List[Dict[str, str]] = []
        for row in ws.iter_rows(min_row=2, values_only=False):
            rid = str(row[0].value) if row and row[0] is not None else ""
            if rid and rid.startswith(prefix):
                out.append({
                    "rid": rid,
                    "row": row[0].row,
                    "date": row[COLS["Data"] - 1].value,
                    "name": row[COLS["Imię"] - 1].value,
                    "place": row[COLS["Miejsce"] - 1].value or "",
                    "start": row[COLS["Start"] - 1].value or "",
                    "end": row[COLS["Koniec"] - 1].value or "",
                    "tasks": row[COLS["Zadania"] - 1].value or "",
                    "notes": row[COLS["Uwagi"] - 1].value or "",
                })
        out.sort(key=lambda e: int(e["rid"].split("_")[-1]))
        return out
    return _with_lock(_read)

def read_entries_all_weeks(user_id: int) -> List[Dict[str, str]]:
    if not os.path.exists(EXCEL_FILE):
        return []
    def _read_all():
        wb = load_workbook(EXCEL_FILE)
        out: List[Dict[str, str]] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if ws.max_row < 2:
                continue
            for row in ws.iter_rows(min_row=2, values_only=False):
                rid = str(row[0].value) if row and row[0] is not None else ""
                if rid and rid.startswith(f"{user_id}_"):
                    out.append({
                        "rid": rid,
                        "date": row[COLS["Data"] - 1].value,
                        "start": row[COLS["Start"] - 1].value or "",
                        "end": row[COLS["Koniec"] - 1].value or "",
                    })
        return out
    return _with_lock(_read_all)

def update_report_field(user_id: int, date_str: str, rid: str, field: str, new_value: str) -> None:
    def _upd():
        wb = load_workbook(EXCEL_FILE)
        ws = ensure_month_sheet(wb, month_key_from_date(date_str))
        col_name_map = {
            "place": "Miejsce",
            "start": "Start",
            "end": "Koniec",
            "tasks": "Zadania",
            "notes": "Uwagi",
        }
        target_col = COLS[col_name_map[field]]
        target_row = None
        for row in ws.iter_rows(min_row=2, values_only=False):
            if str(row[0].value) == rid:
                target_row = row[0].row
                break
        if not target_row:
            raise RuntimeError("Nie znaleziono wiersza do edycji.")
        ws.cell(row=target_row, column=target_col, value=new_value)
        _backup_file()
        _atomic_save_wb(wb, EXCEL_FILE)
    _with_lock(_upd)
    _maybe_upload_sharepoint()

def _maybe_upload_sharepoint() -> None:
    if all([ClientContext, SHAREPOINT_SITE, SHAREPOINT_DOC_LIB, SHAREPOINT_CLIENT_ID, SHAREPOINT_CLIENT_SECRET]):
        try:
            ctx = ClientContext(SHAREPOINT_SITE).with_credentials(
                ClientCredential(SHAREPOINT_CLIENT_ID, SHAREPOINT_CLIENT_SECRET)
            )
            folder = ctx.web.get_folder_by_server_relative_url(SHAREPOINT_DOC_LIB)
            with open(EXCEL_FILE, "rb") as f:
                folder.upload_file(os.path.basename(EXCEL_FILE), f).execute_query()
        except Exception as e:
            logging.warning("SharePoint upload failed: %s", e)

# ──────────────────── presets (miejsca) ────────────────────
def load_mapping() -> Dict[str, int]:
    if os.path.exists(MAPPING_FILE):
        with open(MAPPING_FILE, "r", encoding="utf-8") as f:
            try:
                return json.load(f)
            except Exception:
                return {}
    return {}

def save_mapping(mapping: Dict[str, int]) -> None:
    with open(MAPPING_FILE, "w", encoding="utf-8") as f:
        json.dump(mapping, f)

def load_presets() -> Dict[str, Dict[str, List[str]]]:
    if os.path.exists(PRESETS_FILE):
        with open(PRESETS_FILE, "r", encoding="utf-8") as f:
            try:
                return json.load(f)
            except Exception:
                return {}
    return {}

def save_presets(presets: Dict[str, Dict[str, List[str]]]) -> None:
    with open(PRESETS_FILE, "w", encoding="utf-8") as f:
        json.dump(presets, f, ensure_ascii=False)

def remember_place(user_id: int, place: str) -> None:
    def _upd():
        presets = load_presets()
        key = str(user_id)
        user = presets.setdefault(key, {"places": []})
        if place in user["places"]:
            user["places"].remove(place)
        user["places"].insert(0, place)
        user["places"] = user["places"][:5]
        save_presets(presets)
    _with_lock(_upd)

def get_recent_places(user_id: int) -> List[str]:
    presets = load_presets()
    return presets.get(str(user_id), {}).get("places", [])

# ──────────────────── helpers: time/tags/overlap ────────────────────
def time_to_minutes(t: str) -> int:
    h, m = map(int, t.split(":"))
    return h * 60 + m

def minutes_to_hhmm(m: int) -> str:
    h = m // 60
    mm = m % 60
    return f"{h}h {mm:02d}m"

def extract_tags(text: str) -> List[str]:
    return re.findall(r"#[\wąćęłńóśżźĄĆĘŁŃÓŚŻŹ]+", text or "")

def intervals_overlap(a_start: str, a_end: str, b_start: str, b_end: str) -> bool:
    return max(time_to_minutes(a_start), time_to_minutes(b_start)) < min(time_to_minutes(a_end), time_to_minutes(b_end))

def has_overlap(user_id: int, date_str: str, start: str, end: str, exclude_rid: Optional[str] = None, in_memory: Optional[List[Dict]] = None) -> Tuple[bool, List[Tuple[str, str]]]:
    conflicts = []
    for e in (in_memory or []):
        if e.get("start") and e.get("end") and intervals_overlap(start, end, e["start"], e["end"]):
            conflicts.append((e["start"], e["end"]))
    for e in read_entries_for_day(user_id, date_str):
        if exclude_rid and e["rid"] == exclude_rid:
            continue
        if e["start"] and e["end"] and intervals_overlap(start, end, e["start"], e["end"]):
            conflicts.append((e["start"], e["end"]))
    return (len(conflicts) > 0, conflicts)

def compute_daily_minutes(entries: List[Dict[str, str]]) -> int:
    total = 0
    for e in entries:
        if e.get("start") and e.get("end"):
            total += time_to_minutes(e["end"]) - time_to_minutes(e["start"])
    return total

# ──────────────────── Telegram: sticky + view stack ────────────────────
async def sticky_set(update_or_ctx, context: ContextTypes.DEFAULT_TYPE, text: str, reply_markup: Optional[InlineKeyboardMarkup] = None):
    chat = update_or_ctx.effective_chat if isinstance(update_or_ctx, Update) else None
    chat_id = chat.id if chat else update_or_ctx.callback_query.message.chat.id
    sticky_id = context.user_data.get("sticky_id")
    if sticky_id:
        try:
            await context.bot.edit_message_text(chat_id=chat_id, message_id=sticky_id, text=text, reply_markup=reply_markup)
            return
        except BadRequest as e:
            if "message is not modified" in str(e).lower():
                return
        except Exception:
            pass
    m = await context.bot.send_message(chat_id, text, reply_markup=reply_markup)
    context.user_data["sticky_id"] = m.message_id

async def sticky_delete(context: ContextTypes.DEFAULT_TYPE, chat_id: int):
    sticky_id = context.user_data.get("sticky_id")
    if sticky_id:
        try:
            await context.bot.delete_message(chat_id, sticky_id)
        except Exception:
            pass
        context.user_data.pop("sticky_id", None)

async def safe_answer(q, text: Optional[str] = None, show_alert: bool = False):
    try:
        if text:
            await q.answer(text=text, show_alert=show_alert)
        else:
            await q.answer()
    except BadRequest:
        pass
    except Exception:
        pass

@dataclass
class View:
    name: str
    payload: dict

def push_view(context, name: str, **payload):
    stack = context.user_data.setdefault("view_stack", [])
    stack.append(View(name, payload))

def pop_view(context):
    stack = context.user_data.get("view_stack") or []
    if len(stack) > 1:
        stack.pop()
    return stack[-1] if stack else None

def top_view(context):
    stack = context.user_data.get("view_stack") or []
    return stack[-1] if stack else None

# ──────────────────── Panel renderers ────────────────────
def today_str() -> str:
    return datetime.now().strftime("%d.%m.%Y")

def to_ddmmyyyy(d: date) -> str:
    return d.strftime("%d.%m.%Y")

def format_home(uid: int, date_str: str, name: str) -> str:
    lines = [f"👤 {name} | 📅 {date_str}"]
    exists = report_exists(uid, date_str)
    if exists:
        entries = read_entries_for_day(uid, date_str)
        lines.append("\n📄 *Raport na dziś:*")
        for i, e in enumerate(entries, start=1):
            lines.extend([
                f"#{i} • 📍 {e['place']} • ⏰ {e['start']}-{e['end']}",
                f"📝 {e['tasks'] or '-'}",
                f"💬 {e['notes'] or '-'}",
                ""
            ])
        total = compute_daily_minutes(entries)
        if total:
            lines.append(f"⏳ Suma: {minutes_to_hhmm(total)}")
    else:
        lines.append("\nℹ️ Brak raportu dla tej daty.")
    lines.append("\nWybierz czynność ⬇️")
    return "\n".join(lines)

def build_main_menu(uid: int, date_str: str) -> InlineKeyboardMarkup:
    exists = report_exists(uid, date_str)
    if exists:
        kb = [
            [InlineKeyboardButton(f"📅 Data: {date_str}", callback_data="date:open")],
            [InlineKeyboardButton("✏️ Edytuj raport", callback_data="panel:edit")],
            [InlineKeyboardButton("📥 Eksport", callback_data="export"),
             InlineKeyboardButton("📥 Mój eksport", callback_data="myexport")],
        ]
    else:
        kb = [
            [InlineKeyboardButton(f"📅 Data: {date_str}", callback_data="date:open")],
            [InlineKeyboardButton("📋 Twórz raport", callback_data="panel:create")],
            [InlineKeyboardButton("📥 Eksport", callback_data="export"),
             InlineKeyboardButton("📥 Mój eksport", callback_data="myexport")],
        ]
    return InlineKeyboardMarkup(kb)

def month_kb(year: int, month: int) -> InlineKeyboardMarkup:
    month_name = cal.month_name[month]
    days = cal.monthcalendar(year, month)
    rows = []
    rows.append([InlineKeyboardButton(f"{month_name} {year}", callback_data="noop")])
    rows.append([InlineKeyboardButton(x, callback_data="noop") for x in ["Pn","Wt","Śr","Cz","Pt","So","Nd"]])
    for week in days:
        r = []
        for d in week:
            if d == 0:
                r.append(InlineKeyboardButton(" ", callback_data="noop"))
            else:
                ds = to_ddmmyyyy(date(year, month, d))
                r.append(InlineKeyboardButton(str(d), callback_data=f"day:{ds}"))
        rows.append(r)
    prev_month = (date(year, month, 1) - timedelta(days=1))
    next_month = (date(year, month, cal.monthrange(year, month)[1]) + timedelta(days=1))
    rows.append([
        InlineKeyboardButton("« Poprz", callback_data=f"cal:{prev_month.year}-{prev_month.month:02d}"),
        InlineKeyboardButton("Dziś", callback_data=f"day:{today_str()}"),
        InlineKeyboardButton("Nast »", callback_data=f"cal:{next_month.year}-{next_month.month:02d}"),
    ])
    rows.append([InlineKeyboardButton("↩️ Wstecz", callback_data="nav:back")])
    return InlineKeyboardMarkup(rows)

def placeholder(val: Optional[str]) -> str:
    return val if (val is not None and str(val).strip() != "") else "—"

def panel_create_text(context: ContextTypes.DEFAULT_TYPE) -> str:
    name = context.user_data.get("name", "")
    date_str = context.user_data.get("date", today_str())
    cur = context.user_data.setdefault("current_entry", {})
    entries = context.user_data.get("entries", [])
    aw = context.user_data.get("await") or {}

    dot_tasks = " (●)" if aw.get("field") == "tasks" and aw.get("mode") == "create" else ""
    dot_notes = " (●)" if aw.get("field") == "notes" and aw.get("mode") == "create" else ""

    lines = [
        f"📄 Panel: *Tworzenie raportu*",
        f"👤 Imię: {name}",
        f"📅 Data: {date_str}",
        "",
        f"📍 Miejsce: {placeholder(cur.get('place'))}",
        f"⏰ Start: {placeholder(cur.get('start'))}",
        f"⏰ Koniec: {placeholder(cur.get('end'))}",
        f"📝 Zadania{dot_tasks}:",
        f"{placeholder(cur.get('tasks'))}",
        f"💬 Uwagi{dot_notes}:",
        f"{placeholder(cur.get('notes'))}",
        "",
        f"➕ Pozycje w panelu (niezapisane): {len(entries)}",
    ]
    mins = compute_daily_minutes(entries + ([cur] if cur.get("start") and cur.get("end") else []))
    if mins:
        lines.append(f"⏳ Razem (panel): {minutes_to_hhmm(mins)}")
    lines.append("")
    lines.append("Wybierz czynność ⬇️")
    return "\n".join(lines)

def kb_create(context: ContextTypes.DEFAULT_TYPE) -> InlineKeyboardMarkup:
    back_label = "↩️ Do edycji" if context.user_data.get("from_edit") else "↩️ Wstecz"
    back_cb = "nav:editlist" if context.user_data.get("from_edit") else "nav:home"
    aw = context.user_data.get("await") or {}
    t_lbl = "📝 Zadania (tekst)" + (" ●" if aw.get("field") == "tasks" and aw.get("mode") == "create" else "")
    n_lbl = "💬 Uwagi (tekst)" + (" ●" if aw.get("field") == "notes" and aw.get("mode") == "create" else "")
    kb = [
        [InlineKeyboardButton("📍 Miejsce", callback_data="set:place"),
         InlineKeyboardButton("⏰ Start", callback_data="set:start"),
         InlineKeyboardButton("⏰ Koniec", callback_data="set:end")],
        [InlineKeyboardButton(t_lbl, callback_data="set:tasks"),
         InlineKeyboardButton(n_lbl, callback_data="set:notes")],
        [InlineKeyboardButton("➕ Dodaj pozycję", callback_data="create:add"),
         InlineKeyboardButton("🗑️ Wyczyść pola", callback_data="create:clear")],
        [InlineKeyboardButton("✅ Zakończ raport (zapis do Excela)", callback_data="create:finish")],
        [InlineKeyboardButton(back_label, callback_data=back_cb)],
    ]
    return InlineKeyboardMarkup(kb)

def panel_edit_list_text(context: ContextTypes.DEFAULT_TYPE) -> str:
    date_str = context.user_data.get("date", today_str())
    uid = context.user_data.get("uid")
    entries = read_entries_for_day(uid, date_str)
    context.user_data["edit_entries"] = entries
    lines = [f"✏️ Panel: *Edycja raportu* — {date_str}\n"]
    if not entries:
        lines.append("Brak wpisów dla tej daty.")
    else:
        for i, e in enumerate(entries, start=1):
            lines.extend([
                f"#{i} | 📍 {e['place']} | ⏰ {e['start']}-{e['end']}",
                f"📝 {e['tasks'] or '-'}",
                f"💬 {e['notes'] or '-'}",
                ""
            ])
        total = compute_daily_minutes(entries)
        if total:
            lines.append(f"⏳ Suma: {minutes_to_hhmm(total)}")
    lines.append("\nWybierz pozycję do edycji lub dodaj nową.")
    return "\n".join(lines)

def kb_edit_list(context: ContextTypes.DEFAULT_TYPE) -> InlineKeyboardMarkup:
    entries = context.user_data.get("edit_entries", [])
    rows = []
    for idx, e in enumerate(entries, start=1):
        label = f"#{idx} {e['place']} {e['start']}-{e['end']}"
        rows.append([InlineKeyboardButton(label, callback_data=f"entry:{idx-1}")])
    rows.append([InlineKeyboardButton("➕ Dodaj nową pozycję", callback_data="editlist:addnew")])
    rows.append([InlineKeyboardButton("↩️ Wstecz", callback_data="nav:home")])
    return InlineKeyboardMarkup(rows)

def panel_edit_entry_text(context: ContextTypes.DEFAULT_TYPE) -> str:
    idx = context.user_data.get("edit_idx")
    e = context.user_data.get("edit_entries", [])[idx]
    # kropki dla aktywnego pola w edycji
    aw = context.user_data.get("await") or {}
    dot_tasks = " (●)" if aw.get("field") == "tasks" and aw.get("mode") == "edit" else ""
    dot_notes = " (●)" if aw.get("field") == "notes" and aw.get("mode") == "edit" else ""
    lines = [
        f"✏️ *Edycja pozycji* #{idx+1}",
        f"📍 Miejsce: {e['place']}",
        f"⏰ Start: {e['start']}",
        f"⏰ Koniec: {e['end']}",
        f"📝 Zadania{dot_tasks}:",
        f"{e['tasks'] or '-'}",
        f"💬 Uwagi{dot_notes}:",
        f"{e['notes'] or '-'}",
        "",
        "Co chcesz zmienić?"
    ]
    return "\n".join(lines)

def kb_edit_entry(context: ContextTypes.DEFAULT_TYPE) -> InlineKeyboardMarkup:
    aw = context.user_data.get("await") or {}
    t_lbl = "📝 Zadania" + (" ●" if aw.get("field") == "tasks" and aw.get("mode") == "edit" else "")
    n_lbl = "💬 Uwagi" + (" ●" if aw.get("field") == "notes" and aw.get("mode") == "edit" else "")
    kb = [
        [InlineKeyboardButton("📍 Miejsce", callback_data="editf:place")],
        [InlineKeyboardButton("⏰ Start", callback_data="editf:start"),
         InlineKeyboardButton("⏰ Koniec", callback_data="editf:end")],
        [InlineKeyboardButton(t_lbl, callback_data="editf:tasks"),
         InlineKeyboardButton(n_lbl, callback_data="editf:notes")],
        [InlineKeyboardButton("↩️ Lista pozycji", callback_data="nav:editlist")],
    ]
    return InlineKeyboardMarkup(kb)

def time_kb(selection: dict, back_to: str) -> InlineKeyboardMarkup:
    h = selection.get("h")
    m = selection.get("m")
    rows = []
    for base in [0, 6, 12, 18]:
        row = []
        for x in range(base, min(base+6, 24)):
            mark = "●" if h == x else "○"
            row.append(InlineKeyboardButton(f"{mark} {x:02d}", callback_data=f"t:h:{x:02d}"))
        rows.append(row)
    rowm = []
    for mm in [0, 15, 30, 45]:
        mark = "●" if m == mm else "○"
        rowm.append(InlineKeyboardButton(f"{mark} {mm:02d}", callback_data=f"t:m:{mm:02d}"))
    rows.append(rowm)
    rows.append([InlineKeyboardButton("✅ OK", callback_data="t:ok"),
                 InlineKeyboardButton("❌ Anuluj", callback_data="t:cancel")])
    rows.append([InlineKeyboardButton("↩️ Wstecz", callback_data=f"nav:{back_to}")])
    return InlineKeyboardMarkup(rows)

# ──────────────────── centralny renderer ────────────────────
async def render(update_or_ctx, context: ContextTypes.DEFAULT_TYPE):
    v = top_view(context)
    uid = (update_or_ctx.effective_user.id if isinstance(update_or_ctx, Update)
           else update_or_ctx.callback_query.from_user.id)
    name = context.user_data.get("name", "")
    ds = context.user_data.get("date", today_str())

    if not v or v.name == "home":
        await sticky_set(update_or_ctx, context, format_home(uid, ds, name), build_main_menu(uid, ds))
        return

    if v.name == "calendar":
        y, m = v.payload["year"], v.payload["month"]
        await sticky_set(update_or_ctx, context, "📅 Wybierz datę:", month_kb(y, m))
        return

    if v.name == "create":
        await sticky_set(update_or_ctx, context, panel_create_text(context), kb_create(context))
        return

    if v.name == "edit_list":
        await sticky_set(update_or_ctx, context, panel_edit_list_text(context), kb_edit_list(context))
        return

    if v.name == "edit_entry":
        await sticky_set(update_or_ctx, context, panel_edit_entry_text(context), kb_edit_entry(context))
        return

    if v.name == "place_select_create":
        uid = context.user_data.get("uid")
        await sticky_set(update_or_ctx, context, "📍 Wybierz miejsce:", kb_place_select("create", uid))
        return

    if v.name == "place_select_edit":
        uid = context.user_data.get("uid")
        await sticky_set(update_or_ctx, context, "📍 Wybierz nowe miejsce:", kb_place_select("edit", uid))
        return

    if v.name == "time_pick":
        sel = context.user_data.get("time_edit", {"h": None, "m": 0})
        hh = "--" if sel.get("h") is None else f"{sel['h']:02d}"
        mm = "--" if sel.get("m") is None else f"{sel['m']:02d}"
        title = f"⏰ Ustaw czas (HH:MM)\nWybrane: {hh}:{mm}"
        back_to = "create" if sel.get("mode") == "create" else "editentry"
        await sticky_set(update_or_ctx, context, title, time_kb(sel, back_to=back_to))
        return

def kb_place_select(context_kind: str, uid: int) -> InlineKeyboardMarkup:
    def _rows(user_places: List[str]) -> List[List[InlineKeyboardButton]]:
        rows = []
        for i, p in enumerate(user_places):
            rows.append([InlineKeyboardButton(p, callback_data=f"place_preset:{i}")])
        rows.append([InlineKeyboardButton("✍️ Wpisz ręcznie (wyślij tekst)", callback_data="place_manual")])
        if context_kind == "create":
            rows.append([InlineKeyboardButton("↩️ Wstecz", callback_data="nav:create")])
        else:
            rows.append([InlineKeyboardButton("↩️ Wstecz", callback_data="nav:editentry")])
        return rows
    return InlineKeyboardMarkup(_rows(get_recent_places(uid)))

# ──────────────────── top-level handlers ────────────────────
async def show_menu(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    context.user_data.clear()
    sel_date = today_str()
    context.user_data["date"] = sel_date
    context.user_data["name"] = update.effective_user.first_name
    context.user_data["uid"] = update.effective_user.id
    context.user_data["entries"] = []            # pozycje w panelu (jeszcze nie zapisane do Excela)
    context.user_data["current_entry"] = {}      # edytowana pozycja
    context.user_data["view_stack"] = [View("home", {})]
    await render(update, context)

# ──────────────────── nawigacja i kalendarz ────────────────────
async def main_menu_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await safe_answer(q)
    uid = q.from_user.id
    ds = context.user_data.get("date", today_str())
    if q.data == "date:open":
        now = datetime.now()
        push_view(context, "calendar", year=now.year, month=now.month)
        await render(update, context)
        return
    if q.data == "panel:create":
        if report_exists(uid, ds):
            await safe_answer(q, text="Raport dla tej daty już istnieje. Przechodzę do edycji.", show_alert=False)
            push_view(context, "edit_list")
            await render(update, context)
            return
        context.user_data.pop("from_edit", None)
        push_view(context, "create")
        await render(update, context)
        return
    if q.data == "panel:edit":
        push_view(context, "edit_list")
        await render(update, context)
        return

async def calendar_nav_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await safe_answer(q)
    data = q.data
    if data.startswith("cal:"):
        y, m = map(int, data.split(":")[1].split("-"))
        pop_view(context)
        push_view(context, "calendar", year=y, month=m)
        await render(update, context)
        return DATE_PICK
    elif data.startswith("day:"):
        ds = data.split(":")[1]
        context.user_data["date"] = ds
        # po wyborze daty wracamy do home (z przeglądem raportu dla tej daty)
        context.user_data["view_stack"] = [View("home", {})]
        await render(update, context)
        return ConversationHandler.END
    return DATE_PICK

async def nav_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await safe_answer(q)
    action = q.data.split(":")[1]
    if action == "home":
        context.user_data["view_stack"] = [View("home", {})]
    elif action == "back":
        pop_view(context)
        if not top_view(context):
            context.user_data["view_stack"] = [View("home", {})]
    elif action == "create":
        push_view(context, "create")
    elif action == "editentry":
        push_view(context, "edit_entry")
    elif action == "editlist":
        push_view(context, "edit_list")
    await render(update, context)

# ──────────────────── EXPORT ────────────────────
async def export_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.callback_query:
        await safe_answer(update.callback_query)
    month_arg = None
    if update.callback_query and update.callback_query.data == "export":
        month_arg = month_key_from_date(context.user_data.get("date", today_str()))
    else:
        args = getattr(context, "args", []) or []
        month_arg = args[0] if args else month_key_from_date(today_str())

    if ADMIN_IDS and update.effective_user.id not in ADMIN_IDS:
        await sticky_set(update, context, "Brak uprawnień do eksportu (tylko admini). Użyj /myexport <YYYY-MM>.")
        return ConversationHandler.END

    path = export_month(month_arg)
    if not path:
        await sticky_set(update, context, f"Brak danych dla {month_arg}.")
        return ConversationHandler.END

    with open(path, "rb") as f:
        await update.effective_chat.send_document(f, filename=os.path.basename(path), caption=f"Eksport {month_arg}")
    try:
        os.remove(path)
    except Exception:
        pass
    await render(update, context)
    return ConversationHandler.END

async def myexport_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.callback_query:
        await safe_answer(update.callback_query)

    month_arg = None
    if update.callback_query and update.callback_query.data == "myexport":
        month_arg = month_key_from_date(context.user_data.get("date", today_str()))
    else:
        args = getattr(context, "args", []) or []
        month_arg = args[0] if args else month_key_from_date(today_str())

    path = export_month(month_arg, user_id=update.effective_user.id)
    if not path:
        await sticky_set(update, context, f"Brak danych dla {month_arg}.")
        return ConversationHandler.END

    with open(path, "rb") as f:
        await update.effective_chat.send_document(f, filename=os.path.basename(path), caption=f"Mój eksport {month_arg}")
    try:
        os.remove(path)
    except Exception:
        pass
    await render(update, context)
    return ConversationHandler.END

def export_month(month_key: str, user_id: Optional[int] = None) -> Optional[str]:
    if not os.path.exists(EXCEL_FILE):
        return None
    def _exp() -> Optional[str]:
        wb = load_workbook(EXCEL_FILE)
        if month_key not in wb.sheetnames:
            return None
        ws = wb[month_key]
        out = Workbook()
        wso = out.active
        wso.title = month_key
        wso.append(HEADERS)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            if user_id and not str(row[0]).startswith(f"{user_id}_"):
                continue
            wso.append(list(row))
        tmpf = os.path.join(DATA_DIR, f"export_{month_key}_{user_id or 'ALL'}.xlsx")
        _atomic_save_wb(out, tmpf)
        return tmpf
    return _with_lock(_exp)

# ──────────────────── PANEL: tworzenie wpisów ────────────────────
async def panel_create_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await safe_answer(q)
    data = q.data

    # set:field
    if data.startswith("set:"):
        field = data.split(":")[1]
        if field == "place":
            push_view(context, "place_select_create")
            await render(update, context)
            return
        if field in ("start", "end"):
            cur = context.user_data.setdefault("current_entry", {})
            init_h = None
            init_m = 0
            if cur.get(field):
                try:
                    hh, mm = cur[field].split(":")
                    init_h, init_m = int(hh), int(mm)
                except Exception:
                    init_h, init_m = None, 0
            context.user_data["time_edit"] = {"h": init_h, "m": init_m, "field": field, "mode": "create"}
            push_view(context, "time_pick")
            await render(update, context)
            return
        if field in ("tasks", "notes"):
            context.user_data["await"] = {"mode": "create", "field": field}
            # nie pokazujemy banera u góry – używamy kropek przy przyciskach/sekcjach
            await safe_answer(q, text="Wyślij teraz tekst (wiadomość zostanie usunięta).")
            await render(update, context)
            return

    # wybór miejsca (preset/manual)
    if data.startswith("place_preset:"):
        idx = int(data.split(":")[1])
        places = get_recent_places(context.user_data.get("uid"))
        if idx < len(places):
            context.user_data.setdefault("current_entry", {})["place"] = places[idx]
            await safe_answer(q, text=f"Wybrano: {places[idx]}")
        pop_view(context)
        push_view(context, "create")
        await render(update, context)
        return

    if data == "place_manual":
        context.user_data["await"] = {"mode": "create", "field": "place"}
        await safe_answer(q, text="Wyślij teraz nazwę miejsca (wiadomość zostanie usunięta).")
        await render(update, context)
        return

    # akcje: wyczyść / dodaj / finish
    if data == "create:clear":
        context.user_data["current_entry"] = {}
        context.user_data.pop("await", None)
        await render(update, context)
        return

    if data == "create:add":
        # walidacja + anti-overlap z panelem i Excelem
        cur = context.user_data.get("current_entry", {})
        missing = [k for k in ["place", "start", "end"] if not cur.get(k)]
        if missing:
            await safe_answer(q, text="Uzupełnij: " + ", ".join(missing), show_alert=True)
            return
        if cur["start"] >= cur["end"]:
            await safe_answer(q, text="Start musi być < koniec.", show_alert=True)
            return
        uid = context.user_data.get("uid")
        date_str = context.user_data.get("date", today_str())
        overlap, conflicts = has_overlap(uid, date_str, cur["start"], cur["end"], in_memory=context.user_data.get("entries", []))
        if overlap:
            context.user_data["pending_overlap"] = {"cur": cur, "conflicts": conflicts}
            kb = InlineKeyboardMarkup([
                [InlineKeyboardButton("Kontynuuj mimo to", callback_data="ovl:ok")],
                [InlineKeyboardButton("Zmień godziny", callback_data="ovl:fix")],
                [InlineKeyboardButton("↩️ Wstecz", callback_data="nav:create")],
            ])
            msg = "⚠️ Nakładanie z przedziałami: " + ", ".join([f"{a}-{b}" for a,b in conflicts])
            await sticky_set(update, context, msg, kb)
            return OVERLAP_DECIDE

        context.user_data.setdefault("entries", []).append(cur)
        remember_place(uid, cur["place"])
        context.user_data["current_entry"] = {}
        context.user_data.pop("await", None)
        await safe_answer(q, text="Dodano pozycję.")
        await render(update, context)
        return

    if data.startswith("ovl:"):
        action = data.split(":")[1]
        if action == "ok":
            cur = context.user_data.get("pending_overlap", {}).get("cur")
            if cur:
                context.user_data.setdefault("entries", []).append(cur)
                remember_place(context.user_data.get("uid"), cur["place"])
                context.user_data["current_entry"] = {}
            context.user_data.pop("pending_overlap", None)
            context.user_data.pop("await", None)
            await render(update, context)
            return
        if action == "fix":
            context.user_data.pop("pending_overlap", None)
            context.user_data["time_edit"] = {"h": None, "m": 0, "field": "start", "mode": "create"}
            push_view(context, "time_pick")
            await render(update, context)
            return

    if data == "create:finish":
        # Jeśli bieżąca pozycja kompletna – spróbuj dodać automatycznie (żeby nie wymagać „Dodaj pozycję”)
        cur = context.user_data.get("current_entry", {})
        if cur.get("place") and cur.get("start") and cur.get("end"):
            if cur["start"] >= cur["end"]:
                await safe_answer(q, text="Start musi być < koniec.", show_alert=True)
                return
            uid = context.user_data.get("uid")
            date_str = context.user_data.get("date", today_str())
            overlap, conflicts = has_overlap(uid, date_str, cur["start"], cur["end"], in_memory=context.user_data.get("entries", []))
            if overlap:
                await safe_answer(q, text="Bieżąca pozycja nakłada się czasowo – popraw godziny.", show_alert=True)
                return
            context.user_data.setdefault("entries", []).append(cur)
            remember_place(uid, cur["place"])
            context.user_data["current_entry"] = {}

        entries = context.user_data.get("entries", [])
        if not entries:
            await safe_answer(q, text="Brak pozycji do zapisania.", show_alert=True)
            return

        # Zapis do Excela
        save_report(entries, context.user_data.get("uid"), context.user_data.get("date", today_str()), context.user_data.get("name"))
        context.user_data["entries"] = []
        context.user_data.pop("await", None)
        await safe_answer(q, text="Zapisano raport do Excela.")
        # wróć do ekranu głównego z podglądem raportu
        context.user_data["view_stack"] = [View("home", {})]
        await render(update, context)
        return

# ──────────────────── PANEL: edycja wpisów ────────────────────
async def edit_list_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await safe_answer(q)
    data = q.data
    if data.startswith("entry:"):
        idx = int(data.split(":")[1])
        context.user_data["edit_idx"] = idx
        push_view(context, "edit_entry")
        await render(update, context)
        return
    if data == "editlist:addnew":
        context.user_data["from_edit"] = True
        push_view(context, "create")
        await render(update, context)
        return

async def edit_entry_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await safe_answer(q)
    data = q.data
    idx = context.user_data.get("edit_idx")
    entries = context.user_data.get("edit_entries", [])
    if idx is None or idx >= len(entries):
        await safe_answer(q, text="Pozycja nieznaleziona.", show_alert=True)
        return

    e = entries[idx]
    if data.startswith("editf:"):
        field = data.split(":")[1]
        if field == "place":
            push_view(context, "place_select_edit")
            await render(update, context)
            return
        if field in ("start", "end"):
            init_h = None
            init_m = 0
            try:
                base = e[field]
                if base:
                    hh, mm = base.split(":")
                    init_h, init_m = int(hh), int(mm)
            except Exception:
                pass
            context.user_data["time_edit"] = {"h": init_h, "m": init_m, "field": field, "mode": "edit", "rid": e["rid"]}
            push_view(context, "time_pick")
            await render(update, context)
            return
        if field in ("tasks", "notes"):
            context.user_data["await"] = {"mode": "edit", "field": field, "rid": e["rid"]}
            await safe_answer(q, text="Wyślij teraz nowy tekst (wiadomość zostanie usunięta).")
            await render(update, context)
            return

# ──────────────────── TIME PICKER (create+edit) ────────────────────
async def time_pick_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await safe_answer(q)
    data = q.data
    sel = context.user_data.get("time_edit", {"h": None, "m": 0})
    if data.startswith("t:h:"):
        h = int(data.split(":")[2])
        sel["h"] = h
        context.user_data["time_edit"] = sel
        await render(update, context)
        return
    if data.startswith("t:m:"):
        m = int(data.split(":")[2])
        sel["m"] = m
        context.user_data["time_edit"] = sel
        await render(update, context)
        return
    if data == "t:cancel":
        pop_view(context)
        if sel.get("mode") == "create":
            push_view(context, "create")
        else:
            push_view(context, "edit_entry")
        await render(update, context)
        return
    if data == "t:ok":
        if sel.get("h") is None or sel.get("m") is None:
            await safe_answer(q, text="Wybierz godzinę i minuty.", show_alert=True)
            return
        tval = f"{sel['h']:02d}:{sel['m']:02d}"
        field = sel.get("field")
        mode = sel.get("mode")
        if mode == "create":
            cur = context.user_data.setdefault("current_entry", {})
            cur[field] = tval
            if cur.get("start") and cur.get("end") and cur["start"] >= cur["end"]:
                cur[field] = None
                await safe_answer(q, text="Start musi być < koniec.", show_alert=True)
            pop_view(context)
            push_view(context, "create")
            await render(update, context)
            return
        else:
            rid = sel.get("rid")
            uid = context.user_data.get("uid")
            date_str = context.user_data.get("date", today_str())
            entries = read_entries_for_day(uid, date_str)
            tgt = next((x for x in entries if x["rid"] == rid), None)
            if not tgt:
                await safe_answer(q, text="Pozycja nie istnieje.", show_alert=True)
                pop_view(context); push_view(context, "edit_list")
                await render(update, context)
                return
            new_start = tval if field == "start" else str(tgt["start"]) or tval
            new_end = tval if field == "end" else str(tgt["end"]) or tval
            if new_start and new_end and new_start >= new_end:
                await safe_answer(q, text="Start musi być < koniec.", show_alert=True)
                return
            overlap, conflicts = has_overlap(uid, date_str, new_start, new_end, exclude_rid=rid)
            if overlap:
                await safe_answer(q, text="Godziny nakładają się z innymi wpisami.", show_alert=True)
                return
            try:
                update_report_field(uid, date_str, rid, field, tval)
            except Exception as ex:
                await safe_answer(q, text=f"Błąd zapisu: {ex}", show_alert=True)
            pop_view(context); push_view(context, "edit_list")
            await render(update, context)
            return

# ──────────────────── AWAIT TEXT (create+edit) ────────────────────
async def await_text_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    txt = (update.message.text or "").strip()
    try:
        await update.message.delete()
    except Exception:
        pass

    info = context.user_data.get("await") or {}
    if not info:
        await render(update, context)
        return

    mode = info.get("mode")
    field = info.get("field")

    if mode == "create":
        cur = context.user_data.setdefault("current_entry", {})
        cur[field] = txt
        if field == "place" and txt:
            remember_place(context.user_data.get("uid"), txt)
        context.user_data.pop("await", None)
        if not top_view(context) or top_view(context).name != "create":
            push_view(context, "create")
        await render(update, context)
        return

    if mode == "edit":
        rid = info.get("rid")
        uid = context.user_data.get("uid")
        date_str = context.user_data.get("date", today_str())
        try:
            update_report_field(uid, date_str, rid, field, txt)
        except Exception as ex:
            await sticky_set(update, context, f"❌ Błąd zapisu: {ex}", InlineKeyboardMarkup([[InlineKeyboardButton("↩️ Wstecz", callback_data="nav:editlist")]]))
            context.user_data.pop("await", None)
            return
        context.user_data.pop("await", None)
        if not top_view(context) or top_view(context).name != "edit_list":
            push_view(context, "edit_list")
        await render(update, context)
        return

# ──────────────────── Komendy pomocnicze ────────────────────
async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        await sticky_delete(context, update.effective_chat.id)
    except Exception:
        pass
    await update.effective_chat.send_message("Anulowano.")
    context.user_data.clear()
    return ConversationHandler.END

async def help_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    text = (
        "📘 *Pomoc*\n"
        "• /start – panel główny (z podglądem raportu, jeśli istnieje).\n"
        "• Wprowadzanie tekstów (Miejsce/Zadania/Uwagi): wciśnij przycisk – obok pojawi się kropka (●) – wyślij wiadomość.\n"
        "• Czas ustawiasz przyciskami HH/MM (00 min domyślnie, w edycji pre-selekcja).\n"
        "• Eksport: przyciski lub /export, /myexport.\n"
    )
    await sticky_set(update, context, text, InlineKeyboardMarkup([[InlineKeyboardButton("↩️ Wstecz", callback_data="nav:home")]]))

# ──────────────────── error handler ────────────────────
async def error_handler(update: object, context: ContextTypes.DEFAULT_TYPE) -> None:
    err = context.error
    if isinstance(err, BadRequest) and ("query is not found" in str(err).lower() or "query is too old" in str(err).lower()):
        return
    logging.exception("Unhandled exception: %s", err)

# ──────────────────── PTB Application ────────────────────
async def on_startup(app: Application) -> None:
    await app.bot.set_my_commands([
        BotCommand("start", "Otwórz panel raportów"),
        BotCommand("export", "Eksport (admin): /export YYYY-MM"),
        BotCommand("myexport", "Mój eksport: /myexport YYYY-MM"),
        BotCommand("help", "Pomoc"),
    ])

def build_app() -> Application:
    app = ApplicationBuilder().token(TELEGRAM_TOKEN).post_init(on_startup).build()

    # komendy
    app.add_handler(CommandHandler("start", show_menu))
    app.add_handler(CommandHandler("export", export_handler))
    app.add_handler(CommandHandler("myexport", myexport_handler))
    app.add_handler(CommandHandler("help", help_cmd))

    # top-level menu actions
    app.add_handler(CallbackQueryHandler(main_menu_cb, pattern=r"^(date:open|panel:(create|edit))$"))
    app.add_handler(CallbackQueryHandler(calendar_nav_cb, pattern=r"^(cal:\d{4}-\d{2}|day:\d{2}\.\d{2}\.\d{4})$"))
    app.add_handler(CallbackQueryHandler(nav_handler, pattern=r"^nav:(home|back|create|editentry|editlist)$"))

    # eksporty – osobne callbacki
    app.add_handler(CallbackQueryHandler(export_handler, pattern=r"^export$"))
    app.add_handler(CallbackQueryHandler(myexport_handler, pattern=r"^myexport$"))

    # panel: create
    app.add_handler(CallbackQueryHandler(panel_create_handler, pattern=r"^(set:(place|start|end|tasks|notes)|create:(add|clear|finish)|place_preset:\d+|place_manual|ovl:(ok|fix))$"))

    # panel: edit list / entry
    app.add_handler(CallbackQueryHandler(edit_list_handler, pattern=r"^(entry:\d+|editlist:addnew)$"))
    app.add_handler(CallbackQueryHandler(edit_entry_handler, pattern=r"^editf:(place|start|end|tasks|notes)$"))

    # time picker
    app.add_handler(CallbackQueryHandler(time_pick_handler, pattern=r"^(t:(h|m):\d{2}|t:(ok|cancel))$"))

    # await text – kasujemy i odświeżamy panel
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, await_text_handler))

    # globalny error handler
    app.add_error_handler(error_handler)

    return app

# ──────────────────── main ────────────────────
if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(levelname)-8s | %(name)s | %(message)s")

    if not TELEGRAM_TOKEN:
        raise SystemExit("Brak TELEGRAM_TOKEN w env.")

    bot_app = build_app()

    if WEBHOOK_URL:
        # produkcja – webhook (wymaga: pip install "python-telegram-bot[webhooks]")
        bot_app.run_webhook(
            listen="0.0.0.0",
            port=PORT,
            url_path=TELEGRAM_TOKEN,
            webhook_url=f"{WEBHOOK_URL}/{TELEGRAM_TOKEN}",
        )
    else:
        # lokalnie – polling
        bot_app.run_polling(allowed_updates=Update.ALL_TYPES)
